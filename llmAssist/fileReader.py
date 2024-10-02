import pandas as pd
import subprocess
import olefile
import zlib
import zipfile
import struct
import re
from openpyxl import load_workbook
from tika import parser # Java 설치 필수
from docx import Document
# from docparser import parse# doc실패
#import pypandoc  #doc 실패
# import aspose.words as aw # 저작권 
# import textract # 실패
# from doc2python import reader #실패
# import doc2text #버전 안맞음
import subprocess # libreOffice 설치 필수  어쩔 수 없이 사용
import os
from config import conf


from pdfminer.high_level import extract_text

def read_xls(file_path):
    text = []
    metadata_list = []
    # XLS 파일 읽기 (xls 확장자)
    metadata_data = {"source": file_path, "type": "Document"}
    df = pd.read_excel(file_path, sheet_name=None)
    for sheet_name, data in df.items():
        # text.append(f"Sheet: {sheet_name}")
        # metadata_list.append(metadata_data)
        for row in data.itertuples(index=False):
            row_text = ' '.join(map(str, row))
            text.append(row_text)
            # print(text) 
            
            metadata_list.append(metadata_data)

    return text, metadata_list

def read_xlsx(file_path):
    text = []
    metadata_list = []

    # XLSX 파일 열기
    wb = load_workbook(filename=file_path, read_only=True)
    xlsx_text = []

    # 각 시트의 텍스트 추출
    for sheet in wb.worksheets:
        sheet_text = []
        for row in sheet.iter_rows(values_only=True):
            row_text = [str(cell) for cell in row if cell is not None]
            if row_text:
                sheet_text.append(" | ".join(row_text))
        xlsx_text.append("\n".join(sheet_text))
    
    # 텍스트를 리스트에 추가
    text.append("\n".join(xlsx_text))
    
    # 메타데이터 생성
    metadata_data = {"source": file_path, "type": "Document"}
    metadata_list.append(metadata_data)

    return text, metadata_list

def read_ppt(file_path):
    text = []
    metadata_list = []
    # PPT 파일 읽기
    metadata_data = {"source": file_path, "type": "Document"}
    parsed = parser.from_file(file_path)

    text.append(parsed["content"])
    metadata_list.append(metadata_data)

    return text, metadata_list

def read_pdf(file_path):
    text = []
    metadata_list = []
    
    # PDF 파일에서 텍스트 추출
    pdf_text = extract_text(file_path)

    text.append(pdf_text)
    
    # 메타데이터 생성
    metadata_data = {"source": file_path, "type": "Document"}
    metadata_list.append(metadata_data)

    return text, metadata_list

def read_doc(file_path):
    text = []
    metadata_list = []
    # DOCX 파일 읽기
    output_folder_path = os.path.dirname(file_path)
    
    # 출력 파일 경로 설정
    output_file_path = os.path.splitext(file_path)[0] + '.txt'
    metadata_data = {"source": file_path, "type": "Document"}

    # LibreOffice 명령어로 DOC 파일을 TXT 파일로 변환
    try:
        if conf['os'] == 'window':
            subprocess.run([
                'soffice',
                '--headless',
                '--convert-to', 'txt',
                '--outdir', output_folder_path,
                file_path
            ], check=True)
        elif conf['os'] == 'linux':
            subprocess.run([
                'libreoffice',
                '--headless',
                '--convert-to', 'txt',
                '--outdir', output_folder_path,
                file_path
            ], check=True)

        with open(output_file_path) as file:
            text.append(file.read())
            metadata_list.append(metadata_data)
            
        # 변환된 파일 삭제 (원하는 경우)
        os.remove(output_file_path)
  
        return text, metadata_list
    except subprocess.CalledProcessError as e:
        print(f"Error during conversion: {e}")
        return "", metadata_list
    except FileNotFoundError as e:
        print(f"File not found: {e}")
        return "", metadata_list
    except IOError as e:
        print(f"Error reading or deleting the file: {e}")
        return "", metadata_list    

def read_docx(file_path):
    text = []
    metadata_list = []
    # DOCX 파일 읽기
    metadata_data = {"source": file_path, "type": "Document"}
    doc = Document(file_path)
    for para in doc.paragraphs:
        if para.text.strip():
            text.append(para.text.strip())
            metadata_list.append(metadata_data)
    return text, metadata_list

def read_hwp(file_path):
    text = []
    metadata_list = []
    # HWP 파일 읽기
    metadata_data = {"source": file_path, "type": "Document"}
    f = olefile.OleFileIO(file_path)
    dirs = f.listdir()

    if ["FileHeader"] not in dirs or \
       ["\x05HwpSummaryInformation"] not in dirs:#파일 내부 HwpInfo 와 FileHeader 확인
        raise Exception("Not Valid HWP.")

    # 문서 포맷 압축 여부 확인 (hwp의 내부 데이터는 압축되어 저장 될 경우가 있음)
    header = f.openstream("FileHeader")
    header_data = header.read()
    is_compressed = (header_data[36] & 1) == 1 # 37번째 바이트를 확인하여 최하위 비트가 1인지 확인 -> 압축여부 확인 

    # Body Sections 불러오기 (엑셀 -> sheet)
    nums = []
    for d in dirs:
        if d[0] == "BodyText":
            nums.append(int(d[1][len("Section"):]))
    sections = ["BodyText/Section"+str(x) for x in sorted(nums)]

    # 전체 text 추출
    for section in sections:
        bodytext = f.openstream(section)
        data = bodytext.read()
        if is_compressed:
            unpacked_data = zlib.decompress(data, -15)
        else:
            unpacked_data = data

        # 각 Section 내 text 추출    
        section_text = ""
        i = 0
        size = len(unpacked_data)
        while i < size:
            header = struct.unpack_from("<I", unpacked_data, i)[0]
            rec_type = header & 0x3ff
            rec_len = (header >> 20) & 0xfff

            if rec_type in [67]:
                rec_data = unpacked_data[i+4:i+4+rec_len]
                section_text += rec_data.decode('utf-16')
                section_text += "\n"

            i += 4 + rec_len
        text.append(remove_control_characters(section_text))
        metadata_list.append(metadata_data)
    return text, metadata_list

def read_hwpx(file_path):

    text = []
    metadata_list = []
    metadata_data = {"source": file_path, "type": "Document"}
    with zipfile.ZipFile(file_path, 'r') as zip_ref:
        print (zip_ref)
        with zip_ref.open('Preview/PrvText.txt') as file:
            text.append(file.read().decode('utf-8'))
            metadata_list.append(metadata_data)
    return text, metadata_list


def remove_control_characters(s):
    return re.sub(r'[\u4e00-\u9fff]', '', s)# 한자 삭제
#    return "".join(ch for ch in s if unicodedata.category(ch)[0]!="C") # 컨트롤 문자 제거



